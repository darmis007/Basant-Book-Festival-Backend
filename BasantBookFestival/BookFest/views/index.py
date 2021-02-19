from django.utils import timezone
from django.shortcuts import render, redirect, reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.views.decorators.csrf import csrf_exempt
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from django.core.paginator import Paginator
from django.views.generic import ListView
from django.core import serializers
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Q

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework import status

from celery import shared_task

import os
import re
import datetime
from openpyxl import *
from openpyxl.writer.excel import save_virtual_workbook

from BookFest.models import Buyer, Publisher, Book, Order
from BookFest.helpers.auth_helpers import create_user_from_email, create_publisher_from_email


# Create your views here.


def index(request):
    return HttpResponse('If you can see this, then the backend server is (hopefully) working. \n\t\t\t\t- Darsh Mishra')


def home_index(request):
    return redirect('https://bbf.bits-pilani.ac.in/home/')


@csrf_exempt
def sign_in(request):
    if request.user.is_anonymous:
        return render(request, 'sign_in.html')
    else:
        return redirect('/')


@permission_classes([IsAuthenticated])
@csrf_exempt
def sign_out(request):
    logout(request)
    return Response({
        'message': request.user.username + ' Logged out Successfully'
    })


@api_view(['POST'])
@permission_classes((AllowAny,))
@csrf_exempt
def buyerRegister(request):
    if request.method == 'POST':
        data = request.data
        email = data['email']
        response = {	}
        try:
            buyer = Buyer.objects.get(email=email)
            return Response({'status': 0, 'message': 'Email already exists'})
        except:
            if not re.match(r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)", email):
                return Response({'status': 0, 'message': 'Please enter a valid email address.'})
            try:
                PSRN = data['PSRN']
                name = data['name']
                is_professor = data['is_professor']
                contact = data['contact']
                if 'department' in data.keys():
                    department = data['department']
                else:
                    department = 'Common'
                user = create_user_from_email(
                    name, email, PSRN, is_professor, department)
                try:
                    buyer = Buyer.objects.get(user=user)
                    buyer.contact_no = data['contact']
                    buyer.save()
                except:
                    return Response({'status': 0, 'message': 'Not a valid User Created'})
            except KeyError as missing_data:
                response = Response(
                    {'message': 'Data is Missing: {}'.format(missing_data)})
                return response
            return Response({
                'message': 'User '+'with '+'email '+email+' successfully created'
            })


@api_view(['POST'])
@permission_classes((AllowAny,))
@csrf_exempt
def addBuyerName(request):
    data = request.data
    email = data['email']
    buyer = Buyer.objects.get(email=email)
    buyer.name = data['name']
    buyer.contact_no = data['contact']
    buyer.save()
    return Response({
        'message': 'User '+'with '+'email '+email+' and name '+buyer.name + ' successfully created'
    })


@api_view(['POST'])
@permission_classes((AllowAny,))
@csrf_exempt
def addBookLink(request):
    data = request.data
    isbn = data['ISBN']
    book = Book.objects.filter(ISBN=isbn)
    book.update(image=data['cover'], thumbnail=data['thumbnail'])
    return Response({
        'message': "Book with ISBN:"+isbn+" updated cover image and thumbnail"
    })


@api_view(['POST'])
@permission_classes((AllowAny,))
@csrf_exempt
def publisherRegister(request):
    if request.method == 'POST':
        data = request.data
        email = data['email']
        response = {	}
        try:
            publisher = Publisher.objects.get(email=email)
            return Response({'status': 0, 'message': 'Email already exists'})
        except:
            if not re.match(r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)", email):
                return Response({'status': 0, 'message': 'Please enter a valid email address.'})
            try:
                name = data['name']
                user = create_publisher_from_email(email, name)
                try:
                    publisher = Publisher.objects.get(user=user)
                except:
                    return Response({'status': 0, 'message': 'Not a valid User Created'})
            except KeyError as missing_data:
                response = Response(
                    {'message': 'Data is Missing: {}'.format(missing_data)})
                return response
            return Response({
                'message': 'User '+'with '+'email '+email+' successfully created'
            })


@api_view(['POST'])
@permission_classes((AllowAny,))
def bookRegister(request):
    data = request.data
    try:
        publisher = Publisher.objects.get(name=data['publisher_name'])
    except:
        return Response({
            'message': 'Publisher ' + data['publisher_name'] + ' does not exist'
        })
    try:
        book = Book()
        book.title = data['title']
        book.link = data['link']
        book.publisher = publisher
        book.author = data['author']
        book.edition = data['edition']
        book.year_of_publication = data['year_of_publication']
        book.price_foreign_currency = float(data['price_foreign_currency'])
        book.currency = str(data['currency'])
        book.price_indian_currency = float(data['price_indian_currency'])
        book.ISBN = data['ISBN']
        book.description = data['description']
        book.subject = data['subject']
        book.suply = data['supply']
        book.discount = int(data['discount'])
        book.image = "http://covers.openlibrary.org/b/isbn/" + \
            data['ISBN']+"-M.jpg"
        book.thumbnail = "http://covers.openlibrary.org/b/isbn/" + \
            data['ISBN']+"-S.jpg"
        #book.expected_price = data['expected_price']
        book.save()
    except KeyError as missing_data:
        response = Response(
            {'message': 'Data is Missing: {}'.format(missing_data)})
        return response
    return Response({
        'data': book.to_dict(),
        'message': 'book created successfully'
    }, status=status.HTTP_201_CREATED,
    )


@api_view(['GET'])
@permission_classes((AllowAny,))
def getBooks(request, page_number):
    listBooks = Paginator(Book.objects.values(), 10)
    page = listBooks.page(page_number)
    return JsonResponse({'data': list(page)})


@api_view(['GET'])
@permission_classes((AllowAny,))
def getPublishers(request):
    publishers = Publisher.objects.all()
    return Response({
        'data': publishers.values()
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes((AllowAny,))
def getAllBooks(request):
    books = Book.objects.all()
    return Response({
        'data': books.values()
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes((AllowAny,))
@shared_task
def getAllSubjects(request):
    subjects = ['Humanities and Social Science', 'Mathematics', 'Management', 'Mechanical Engineering', 'Pharmacy', 'Physics', 'Computer Science', 'Electrical and Electronic Engineering',
                'Chemical Engineering', 'Physics', 'Bio Science', 'Economics & Finance', 'General Reading', 'Biographies', 'Fictions', 'Civil Engineering', 'Chemistry']
    return Response({
        'data': subjects
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes((AllowAny,))
def getBook(request, book_id):
    try:
        book = Book.objects.filter(id=book_id)
    except:
        return Response({
            'message': "Book Id "+str(book_id) + " does not exist"
        })
    #dat = serializers.serialize('json', book)
    return JsonResponse({'data': list(book.values())})


@api_view(['POST'])
@permission_classes((AllowAny,))
@csrf_exempt
@shared_task
def filterBooks(request, search_type):
    types = ["title", "author", "description", "subject", "publisher"]
    if search_type == None or search_type not in types:
        return Response({
            "message": "not a valid search_type"
        }, status=status.HTTP_400_BAD_REQUEST)
    else:
        try:
            data = request.data
            if search_type == "author":
                author_data = data['search']
                books = Book.objects.filter(author__icontains=author_data)
            elif search_type == "description":
                description_data = data['search']
                books = Book.objects.filter(
                    description__icontains=description_data)
            elif search_type == "title":
                title_data = data['search']
                books = Book.objects.filter(title__icontains=title_data)
            elif search_type == "subject":
                subject_data = data['search']
                books = Book.objects.filter(subject__icontains=subject_data)
            elif search_type == "publisher":
                publisher_data = data["search"]
                books = Book.objects.filter(publisher__id=int(publisher_data))
            return JsonResponse({'data': list(books.values())})
        except Exception as e:
            return Response({
                "message": e
            })


@api_view(['POST'])
@permission_classes((AllowAny,))
@csrf_exempt
@shared_task
def filterPublisherSubjectBooks(request, publisher, search_type):
    types = ['Humanities and Social Science', 'Mathematics', 'Management', 'Mechanical Engineering', 'Pharmacy', 'Physics', 'Computer Science', 'Electrical and Electronic Engineering',
             'Chemical Engineering', 'Physics', 'Bio Science', 'Economics & Finance', 'General Reading', 'Biographies', 'Fictions', 'Civil Engineering', 'Chemistry']

    if search_type == None or search_type not in types:
        return Response({
            "message": "not a valid search_type"
        }, status=status.HTTP_400_BAD_REQUEST)
    else:
        try:
            books = Book.objects.filter(
                publisher__id=publisher, subject__icontains=search_type)
            return JsonResponse({'data': list(books.values())})
        except Exception as e:
            return Response({
                "message": e
            })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@csrf_exempt
def placeOrder(request):
    data = request.data
    try:
        book_id = data['book_id']
        book = Book.objects.get(id=book_id)
    except KeyError as missing_key:
        return Response({
            'message': missing_key+' id not sent'
        }, status=status.HTTP_400_BAD_REQUEST)
    except ObjectDoesNotExist:
        return Response({
            'message': 'Book is not available'
        }, status=status.HTTP_400_BAD_REQUEST)
    try:
        order = Order()
        buyer = Buyer.objects.get(user=request.user)
        publisher = book.publisher
        order.buyer = buyer
        order.seller = publisher
        order.book = book
        if request.user.buyer_profile.is_professor == False:
            order.recommended_to_library = True
        else:
            if data['recommended'] == "false" or data['recommended'] == "False":
                order.recommended_to_library = False
            else:
                order.recommended_to_library = True
        order.is_ordered = True
        order.save()
        return Response({
            'message': "Order "+str(order.id)+" Placed Successfully for "+book.title+" by "+book.publisher.name,
            'order': order.id
        })
    except Exception as e:
        return Response({
            'message': e
        })


'''
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def placeOrder(request):
    data = request.data
    try:
        book_id = data['book_id']
        book = Book.objects.get(id=book_id)
    except KeyError:
        return Response({
            'message':'Book id not sent'
        })
    except ObjectDoesNotExist:
        return Response({
            'message':'Book is not available'
        })

    try:
        order = Order.objects.get(book__id=book, buyer__user=request.user)
        return Response({
            'message': "You have already placed an order for "+book.title+" "+book.publisher.name+" with order ID "+order.id
        })

    except ObjectDoesNotExist:
        try:
            order = Order()
            buyer = Buyer.objects.get(user=request.user)
            publisher = book.publisher
            order.buyer = buyer
            order.seller = publisher
            order.book = book
            if request.user.buyer_profile.is_professor == False :
                order.recommended_to_library = True
            else:
                order.recommended_to_library = data['recommended']
            order.is_ordered = True
            order.save()
            return Response({
                'message':"Order "+order.id+" Placed Successfully for "+book.title+" by "+book.publisher.name,
                'order':order
            })
        except Exception as e:
        return Response({
            'message': e
        })
'''


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@csrf_exempt
def cancelOrder(request):
    data = request.data
    try:
        order_id = data['order_id']
    except KeyError:
        return Response({
            'message': "Order Id not sent"
        }, status=status.HTTP_400_BAD_REQUEST)
    try:
        order = Order.objects.get(id=order_id, is_ordered=True)
    except ObjectDoesNotExist:
        return Response({
            'message': "Order ID "+str(order.id)+" does not exist or has cancelled"
        }, status=status.HTTP_400_BAD_REQUEST)
    if request.user != order.buyer.user:
        return Response({
            'message': "You are not allowed to perform the operation"
        }, status=status.HTTP_400_BAD_REQUEST)
    order.is_ordered = False
    order.save()
    return Response({
        'message': 'Order '+str(order.id)+' cancelled succesfully by you'
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@csrf_exempt
def myOrders(request):
    try:
        library_orders = Order.objects.filter(
            is_ordered=True, buyer__user=request.user, recommended_to_library=True)
        cancelled_orders = Order.objects.filter(
            is_ordered=False, buyer__user=request.user)
        personal_orders = Order.objects.filter(
            is_ordered=True, buyer__user=request.user, recommended_to_library=False)
        return Response({
            'library': library_orders.values(),
            'cancelled': cancelled_orders.values(),
            'personal': personal_orders.values(),
        }, status=status.HTTP_200_OK
        )
    except Exception as e:
        return Response({
            'message': e + " has occured."
        }, status=status.HTTP_400_BAD_REQUEST)


@staff_member_required
def orderedExcel(request):
    orders = Order.objects.filter(is_ordered=True)
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Order ID"
    ws["B1"] = "Buyer Name"
    ws["C1"] = "Buyer Email"
    ws["D1"] = "Buyer PSRN"
    ws["E1"] = "Book"
    ws["F1"] = "Publisher"
    ws["G1"] = "Recommended to Library"
    ws["H1"] = "Book Price"
    ws["I1"] = "Order Time"
    ws["J1"] = "Buyer Department"
    ws["K1"] = "Subject"
    ws["L1"] = "Book ISBN"
    ws["M1"] = "BBF Book ID"
    ws["N1"] = "Book Author"
    ws["O1"] = "Book Year of Publication"
    row = 2
    for order in orders:
        ws["A{}".format(row)] = order.id
        ws["B{}".format(row)] = order.buyer.name
        ws["C{}".format(row)] = order.buyer.email
        ws["D{}".format(row)] = order.buyer.PSRN
        ws["E{}".format(row)] = order.book.title
        ws["F{}".format(row)] = order.seller.name
        ws["G{}".format(row)] = order.recommended_to_library
        ws["H{}".format(row)] = round(order.book.expected_price, 0)
        ws["I{}".format(row)] = order.created_at
        ws["J{}".format(row)] = order.buyer.department
        ws["K{}".format(row)] = order.book.subject
        ws["L{}".format(row)] = order.book.ISBN
        ws["M{}".format(row)] = order.book.id
        ws["N{}".format(row)] = order.book.author
        ws["O{}".format(row)] = order.book.year_of_publication

        row += 1

    response = HttpResponse(content=save_virtual_workbook(
        wb), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=Master_List_All_Orders.xlsx"
    return response


@staff_member_required
def orderedPublisherExcel(request, publisher_id):
    orders = Order.objects.filter(is_ordered=True, seller__id=publisher_id)
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Order ID"
    ws["B1"] = "Buyer Name"
    ws["C1"] = "Buyer Email"
    ws["D1"] = "Buyer PSRN"
    ws["E1"] = "Book"
    ws["F1"] = "Publisher"
    ws["G1"] = "Recommended to Library"
    ws["H1"] = "Book Price"
    ws["I1"] = "Order Time"
    ws["J1"] = "Buyer Department"
    ws["K1"] = "Book Department"
    ws["L1"] = "Book ISBN"
    ws["M1"] = "Book ID"
    ws["N1"] = "Book Author"
    ws["O1"] = "Book Year of Publication"
    row = 2
    for order in orders:
        ws["A{}".format(row)] = order.id
        ws["B{}".format(row)] = order.buyer.name
        ws["C{}".format(row)] = order.buyer.email
        ws["D{}".format(row)] = order.buyer.PSRN
        ws["E{}".format(row)] = order.book.title
        ws["F{}".format(row)] = order.seller.name
        ws["G{}".format(row)] = order.recommended_to_library
        ws["H{}".format(row)] = order.book.expected_price
        ws["I{}".format(row)] = order.created_at
        ws["J{}".format(row)] = order.buyer.department
        ws["K{}".format(row)] = order.book.subject
        ws["L{}".format(row)] = order.book.ISBN
        ws["M{}".format(row)] = order.book.id
        ws["N{}".format(row)] = order.book.author
        ws["O{}".format(row)] = order.book.year_of_publication

        row += 1

    response = HttpResponse(content=save_virtual_workbook(
        wb), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename= {} Master_List_All_Orders.xlsx".format(
        order.seller.name)
    return response


def booksExcel(request):
    books = Book.objects.all()
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "BBF Book ID"
    ws["B1"] = "Title"
    ws["C1"] = "Author"
    ws["D1"] = "Edition"
    ws["E1"] = "Publisher"
    ws["F1"] = "ISBN"
    ws["G1"] = "Price (After Discount)"
    ws["H1"] = "Subject"
    ws["I1"] = "Year of Publication"
    ws["J1"] = "Book Link"

    row = 2
    for book in books:
        ws["A{}".format(row)] = book.id
        ws["B{}".format(row)] = book.title
        ws["C{}".format(row)] = book.author
        ws["D{}".format(row)] = book.edition
        ws["E{}".format(row)] = book.publisher.name
        ws["F{}".format(row)] = book.ISBN
        ws["G{}".format(row)] = round(book.expected_price, 0)
        ws["H{}".format(row)] = book.subject
        ws["I{}".format(row)] = book.year_of_publication
        ws["J{}".format(row)] = book.link
        print(row)
        row += 1

    response = HttpResponse(content=save_virtual_workbook(
        wb), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=Master_List_All_Books.xlsx"

    print("File Created.")
    return response
